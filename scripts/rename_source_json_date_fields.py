#!/usr/bin/env python3
"""Rename and pair date fields in 奏摺上諭結構化數據.json.

Rules:
- memorial_date -> send_date
- rescript_date -> receive_date
- for 上諭: issue_date -> announce_date
- for other doc_type values: keep issue_date name
- every real date value becomes [Chinese date, Arabic date]

Arabic dates are read from the workbook:
- send_date: 上奏日期_西歷, fallback 西歷日期
- receive_date: 硃批日期_西歷, fallback 西歷日期
- announce_date / issue_date: 西歷日期

No online date conversion is used.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_JSON = Path("/Users/creamybanana/Downloads/林爽文/Final Drive Copy/奏摺上諭結構化數據.json")
DEFAULT_XLSM = Path("/Users/creamybanana/Downloads/林爽文/林爽文Final/（3）原始史料統計/奏摺上諭名單.xlsm")


CN_NUM = {
    "〇": 0,
    "零": 0,
    "一": 1,
    "二": 2,
    "兩": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


def normalize_arabic(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.year:04d}/{value.month:02d}/{value.day:02d}"
    if isinstance(value, date):
        return f"{value.year:04d}/{value.month:02d}/{value.day:02d}"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def cn_int(text: str | None) -> int | None:
    if not text:
        return None
    text = text.strip().replace("初", "").replace("廿", "二十").replace("卅", "三十")
    if text in {"正", "元"}:
        return 1
    if text == "冬":
        return 11
    if text in {"臘", "腊"}:
        return 12
    if "十" in text:
        left, _, right = text.partition("十")
        tens = CN_NUM.get(left, 1) if left else 1
        ones = CN_NUM.get(right, 0) if right else 0
        return tens * 10 + ones
    value = 0
    for ch in text:
        if ch not in CN_NUM:
            return None
        value = value * 10 + CN_NUM[ch]
    return value if value else None


DATE_RE = re.compile(
    r"乾隆(?P<year>[〇零一二兩两三四五六七八九十]+)年"
    r"(?:(?P<month>正|元|冬|臘|腊|[〇零一二兩两三四五六七八九十]+)月)?"
    r"(?:(?P<day>初?[〇零一二兩两三四五六七八九十廿卅]+)日)?"
)


def fallback_arabic_from_chinese(chinese: Any) -> str | None:
    if not isinstance(chinese, str):
        return None
    if chinese.strip() in {"", "/"}:
        return None
    m = DATE_RE.search(chinese)
    if not m:
        return None
    year = cn_int(m.group("year"))
    if year is None:
        return None
    western_year = year + 1735
    month = cn_int(m.group("month"))
    day = cn_int(m.group("day"))
    if month is None:
        return str(western_year)
    if day is None:
        return f"{western_year:04d}/{month:02d}"
    return f"{western_year:04d}/{month:02d}/{day:02d}"


def date_pair(chinese: Any, arabic: Any) -> Any:
    if chinese is None:
        return None
    if isinstance(chinese, str) and chinese.strip() == "":
        return chinese
    arabic_norm = normalize_arabic(arabic) or fallback_arabic_from_chinese(chinese)
    return [chinese, arabic_norm]


def load_workbook_lookup(xlsm_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(xlsm_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb["All"]
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {header: i for i, header in enumerate(headers) if header}
    lookup: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        doc_id = row[idx["id"]]
        if not doc_id:
            continue
        lookup[str(doc_id)] = {
            "western_title_date": row[idx["西歷日期"]],
            "western_memorial_date": row[idx["上奏日期_西歷"]],
            "western_rescript_date": row[idx["硃批日期_西歷"]],
        }
    return lookup


def transform_record(record: dict[str, Any], lookup: dict[str, dict[str, Any]]) -> dict[str, Any]:
    out = dict(record)
    doc_id = str(out.get("doc_id"))
    dates = lookup.get(doc_id, {})

    if "memorial_date" in out:
        out["send_date"] = date_pair(
            out.pop("memorial_date"),
            dates.get("western_memorial_date") or dates.get("western_title_date"),
        )

    if "rescript_date" in out:
        out["receive_date"] = date_pair(
            out.pop("rescript_date"),
            dates.get("western_rescript_date") or dates.get("western_title_date"),
        )

    if "issue_date" in out:
        new_name = "announce_date" if out.get("doc_type") == "上諭" else "issue_date"
        value = out.pop("issue_date")
        out[new_name] = date_pair(value, dates.get("western_title_date"))

    return out


def atomic_write_json(path: Path, data: list[dict[str, Any]]) -> None:
    text = json.dumps(data, ensure_ascii=False, indent=2)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(text)
            f.write("\n")
        os.replace(tmp_name, path)
    except Exception:
        try:
            os.unlink(tmp_name)
        except FileNotFoundError:
            pass
        raise


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--xlsm", type=Path, default=DEFAULT_XLSM)
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args()

    lookup = load_workbook_lookup(args.xlsm)
    data = json.loads(args.json.read_text(encoding="utf-8"))
    transformed = [transform_record(row, lookup) for row in data]

    summary = {
        "records": len(data),
        "workbook_lookup_records": len(lookup),
        "old_date_field_counts_after_transform": {
            "memorial_date": sum(1 for row in transformed if "memorial_date" in row),
            "rescript_date": sum(1 for row in transformed if "rescript_date" in row),
            "issue_date_on_shangyu": sum(1 for row in transformed if row.get("doc_type") == "上諭" and "issue_date" in row),
        },
        "new_date_field_counts": {
            "send_date": sum(1 for row in transformed if "send_date" in row),
            "receive_date": sum(1 for row in transformed if "receive_date" in row),
            "announce_date": sum(1 for row in transformed if "announce_date" in row),
            "issue_date": sum(1 for row in transformed if "issue_date" in row),
        },
        "date_fields_by_type": {
            field: dict(Counter(row.get("doc_type") for row in transformed if field in row))
            for field in ("send_date", "receive_date", "announce_date", "issue_date")
        },
        "bad_date_pair_values": sum(
            1
            for row in transformed
            for field in ("send_date", "receive_date", "announce_date", "issue_date")
            if field in row and row[field] is not None and not isinstance(row[field], list)
        ),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.write:
        atomic_write_json(args.json, transformed)
        print(f"wrote {args.json}")
    else:
        print("dry run only; pass --write to update the JSON")


if __name__ == "__main__":
    main()
