#!/usr/bin/env python3
"""Update 上諭 dates in 奏摺上諭結構化數據.json from the xlsm title/date list.

Rules:
- Match JSON records where doc_type == "上諭" by title.
- Use the workbook row where 類別 == "上諭" and 標題 matches.
- Read 西歷日期 from the workbook.
- Convert the western year/month/day mechanically to a Qianlong Chinese date:
  1786 -> 乾隆五十一年, 1787 -> 乾隆五十二年, etc.
- Put the converted Chinese date in issue_date.
- Remove memorial_date and rescript_date from 上諭 records.
- Do not use online date conversion.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_JSON = Path("/Users/creamybanana/Downloads/林爽文/Final Drive Copy/奏摺上諭結構化數據.json")
DEFAULT_XLSM = Path("/Users/creamybanana/Downloads/林爽文/林爽文Final/（3）原始史料統計/奏摺上諭名單.xlsm")


def int_to_chinese(n: int) -> str:
    digits = "零一二三四五六七八九"
    if n <= 0:
        raise ValueError(f"cannot convert non-positive integer: {n}")
    if n < 10:
        return digits[n]
    if n == 10:
        return "十"
    if n < 20:
        return "十" + digits[n % 10]
    if n < 100:
        tens, ones = divmod(n, 10)
        return digits[tens] + "十" + (digits[ones] if ones else "")
    raise ValueError(f"integer too large for this converter: {n}")


def parse_western_date(value: Any) -> tuple[int, int | None, int | None]:
    if isinstance(value, datetime):
        return value.year, value.month, value.day
    if isinstance(value, date):
        return value.year, value.month, value.day
    if isinstance(value, int):
        return value, None, None
    if isinstance(value, float) and value.is_integer():
        return int(value), None, None
    if not isinstance(value, str):
        raise ValueError(f"unsupported date value: {value!r}")

    text = value.strip()
    m = re.fullmatch(r"(\d{4})(?:/(\d{1,2})(?:/(\d{1,2}))?)?", text)
    if not m:
        raise ValueError(f"unsupported western date string: {value!r}")
    year = int(m.group(1))
    month = int(m.group(2)) if m.group(2) else None
    day = int(m.group(3)) if m.group(3) else None
    return year, month, day


def western_to_qianlong_chinese(value: Any) -> str:
    year, month, day = parse_western_date(value)
    qianlong_year = year - 1735
    result = f"乾隆{int_to_chinese(qianlong_year)}年"
    if month is not None:
        result += f"{int_to_chinese(month)}月"
    if day is not None:
        result += f"{int_to_chinese(day)}日"
    return result


def load_title_dates(xlsm_path: Path) -> dict[str, str]:
    wb = load_workbook(xlsm_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb["All"]
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: i for i, header in enumerate(headers) if header}
    required = {"類別", "標題", "西歷日期"}
    missing = required - set(index)
    if missing:
        raise RuntimeError(f"missing workbook columns: {sorted(missing)}")

    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[index["類別"]] != "上諭":
            continue
        title = row[index["標題"]]
        western_date = row[index["西歷日期"]]
        if not title:
            continue
        converted = western_to_qianlong_chinese(western_date)
        if title in lookup and lookup[title] != converted:
            raise RuntimeError(f"duplicate title with conflicting date: {title}")
        lookup[title] = converted
    return lookup


def update_json(data: list[dict[str, Any]], title_dates: dict[str, str]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    updated: list[dict[str, Any]] = []
    missing_titles: list[str] = []
    changed = 0
    for record in data:
        out = dict(record)
        if out.get("doc_type") == "上諭":
            title = out.get("title")
            if title not in title_dates:
                missing_titles.append(title)
            else:
                new_issue = title_dates[title]
                if out.get("issue_date") != new_issue or "memorial_date" in out or "rescript_date" in out:
                    changed += 1
                out["issue_date"] = new_issue
                out.pop("memorial_date", None)
                out.pop("rescript_date", None)
        updated.append(out)

    summary = {
        "shangyu_lookup_titles": len(title_dates),
        "shangyu_records": sum(1 for row in data if row.get("doc_type") == "上諭"),
        "changed_shangyu_records": changed,
        "missing_titles": missing_titles,
    }
    return updated, summary


def atomic_write_json(path: Path, data: list[dict[str, Any]]) -> None:
    text = json.dumps(data, ensure_ascii=False, indent=2)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(text)
            f.write("\n")
        os.replace(tmp_name, path)
    except Exception:
        try:
            os.unlink(tmp_name)
        except FileNotFoundError:
            pass
        raise


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--xlsm", type=Path, default=DEFAULT_XLSM)
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args()

    title_dates = load_title_dates(args.xlsm)
    data = json.loads(args.json.read_text(encoding="utf-8"))
    updated, summary = update_json(data, title_dates)

    if summary["missing_titles"]:
        raise RuntimeError(f"missing titles: {summary['missing_titles'][:10]}")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if args.write:
        atomic_write_json(args.json, updated)
        print(f"wrote {args.json}")
    else:
        print("dry run only; pass --write to update the JSON")


if __name__ == "__main__":
    main()
